import os
import sys
import shutil
import tkinter as tk
from tkinter import ttk
from PIL import Image, ImageTk
from openpyxl import load_workbook
from openpyxl.utils import range_boundaries


APP_TITLE = "Orçamento System"
WINDOW_SIZE = "920x520"
EXCEL_FILE = "interface.xlsx"
RUNTIME_EXCEL_FILE = "interface_runtime.xlsx"
LOGO_FILE = "logo.jpg"


def resource_path(relative_path: str) -> str:
    base_path = getattr(sys, "_MEIPASS", os.path.abspath(os.path.dirname(__file__)))
    return os.path.join(base_path, relative_path)


def format_brl(value) -> str:
    if value is None:
        value = 0
    try:
        number = float(value)
    except (TypeError, ValueError):
        number = 0.0
    formatted = f"{number:,.2f}"
    formatted = formatted.replace(",", "X").replace(".", ",").replace("X", ".")
    return f"R$ {formatted}"


def parse_float(text: str) -> float:
    if text is None:
        return 0.0
    cleaned = str(text).strip().replace(".", "").replace(",", ".")
    try:
        return float(cleaned)
    except ValueError:
        return 0.0


def parse_int(text: str) -> int:
    if text is None:
        return 0
    cleaned = str(text).strip()
    try:
        return int(cleaned)
    except ValueError:
        return 0


def cell_in_sqref(cell: str, sqref) -> bool:
    for rng in sqref.ranges:
        if cell in rng:
            return True
    return False


def normalize_sheet_name(name: str) -> str:
    name = name.strip()
    if name.startswith("'") and name.endswith("'"):
        return name[1:-1]
    return name


def values_from_range(ws, cell_range: str):
    values = []
    try:
        min_col, min_row, max_col, max_row = range_boundaries(cell_range)
    except ValueError:
        return values
    for row in ws.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col):
        for cell in row:
            if cell.value is not None:
                values.append(str(cell.value))
    return values


def get_dropdown_values(ws, cell: str):
    validations = getattr(ws, "data_validations", None)
    if not validations:
        return []
    for dv in validations.dataValidation:
        if not cell_in_sqref(cell, dv.sqref):
            continue
        formula = dv.formula1
        if not formula:
            return []
        formula = str(formula).strip()
        if formula.startswith("="):
            formula = formula[1:]
        if "," in formula and not "!" in formula and not ":" in formula and not formula.startswith("$"):
            cleaned = formula.strip().strip('"')
            return [item.strip() for item in cleaned.split(",") if item.strip()]
        if "!" in formula:
            sheet_name, cell_range = formula.split("!", 1)
            sheet_name = normalize_sheet_name(sheet_name)
            target_ws = ws.parent[sheet_name] if sheet_name in ws.parent.sheetnames else ws
            return values_from_range(target_ws, cell_range.replace("$", ""))
        if ":" in formula:
            return values_from_range(ws, formula.replace("$", ""))
        cleaned = formula.strip().strip('"')
        if "," in cleaned:
            return [item.strip() for item in cleaned.split(",") if item.strip()]
        return [cleaned] if cleaned else []
    return []


class OrcamentoApp:
    def __init__(self, root: tk.Tk):
        self.root = root
        self.root.title(APP_TITLE)
        self.root.geometry(WINDOW_SIZE)
        self.root.resizable(False, False)

        self.excel_path = resource_path(EXCEL_FILE)
        self.runtime_excel_path = resource_path(RUNTIME_EXCEL_FILE)
        self.logo_path = resource_path(LOGO_FILE)

        self.dropdown_cache = {}
        self._load_dropdowns()
        self._build_ui()

    def _load_dropdowns(self):
        try:
            wb = load_workbook(self.excel_path)
        except FileNotFoundError:
            self.dropdown_cache = {}
            return
        ws = wb.active
        mapping = {
            "B3": "contribuinte",
            "B7": "pagamento",
            "B8": "bandeira",
            "B9": "parcelas",
            "B10": "estado",
        }
        for cell, key in mapping.items():
            self.dropdown_cache[key] = get_dropdown_values(ws, cell)
        wb.close()

    def _build_ui(self):
        container = ttk.Frame(self.root, padding=10)
        container.pack(fill="both", expand=True)

        left_frame = ttk.Frame(container)
        left_frame.pack(side="left", fill="both", expand=True)

        right_frame = ttk.Frame(container)
        right_frame.pack(side="right", fill="both", expand=True)

        logo_frame = ttk.Frame(left_frame)
        logo_frame.pack(anchor="w")
        self._load_logo(logo_frame)

        form_frame = ttk.Frame(left_frame)
        form_frame.pack(fill="both", expand=True, pady=(10, 0))

        self.contribuinte_var = tk.StringVar()
        self.pagamento_var = tk.StringVar()
        self.bandeira_var = tk.StringVar()
        self.parcelas_var = tk.StringVar()
        self.estado_var = tk.StringVar()
        self.preco_var = tk.StringVar()
        self.quantidade_var = tk.StringVar()

        self._add_label(form_frame, "Contribuinte", 0)
        self.contribuinte_combo = self._add_combo(form_frame, self.contribuinte_var, "contribuinte", 0)

        self._add_label(form_frame, "Preço", 1)
        self.preco_entry = self._add_entry(form_frame, self.preco_var, 1)

        self._add_label(form_frame, "Quantidade", 2)
        self.quantidade_entry = self._add_entry(form_frame, self.quantidade_var, 2)

        self._add_label(form_frame, "Pagamento", 3)
        self.pagamento_combo = self._add_combo(form_frame, self.pagamento_var, "pagamento", 3)
        self.pagamento_combo.bind("<<ComboboxSelected>>", self._update_pagamento_state)

        self._add_label(form_frame, "Bandeira", 4)
        self.bandeira_combo = self._add_combo(form_frame, self.bandeira_var, "bandeira", 4)

        self._add_label(form_frame, "Parcelas", 5)
        self.parcelas_combo = self._add_combo(form_frame, self.parcelas_var, "parcelas", 5)

        self._add_label(form_frame, "Estado", 6)
        self.estado_combo = self._add_combo(form_frame, self.estado_var, "estado", 6)

        button_frame = ttk.Frame(form_frame)
        button_frame.grid(row=7, column=0, columnspan=2, pady=10, sticky="w")
        ttk.Button(button_frame, text="Calcular", command=self.calculate).pack(anchor="w")

        result_frame = ttk.LabelFrame(right_frame, text="Resultados", padding=10)
        result_frame.pack(fill="both", expand=True)

        self.frete_var = tk.StringVar(value=format_brl(0))
        self.default_var = tk.StringVar(value=format_brl(0))
        self.total_var = tk.StringVar(value=format_brl(0))
        self.valor_unit_var = tk.StringVar(value=format_brl(0))

        self._add_result(result_frame, "Frete", self.frete_var, 0)
        self._add_result(result_frame, "Default", self.default_var, 1)
        self._add_result(result_frame, "Total Cliente", self.total_var, 2)
        self._add_result(result_frame, "Valor Unitário", self.valor_unit_var, 3)

        self._update_pagamento_state()

    def _load_logo(self, frame: ttk.Frame):
        if not os.path.exists(self.logo_path):
            return
        try:
            image = Image.open(self.logo_path)
            image.thumbnail((160, 160))
            self.logo_image = ImageTk.PhotoImage(image)
            ttk.Label(frame, image=self.logo_image).pack(anchor="w")
        except Exception:
            pass

    @staticmethod
    def _add_label(frame: ttk.Frame, text: str, row: int):
        ttk.Label(frame, text=text).grid(row=row, column=0, sticky="w", pady=4, padx=(0, 8))

    def _add_entry(self, frame: ttk.Frame, variable: tk.StringVar, row: int):
        entry = ttk.Entry(frame, textvariable=variable, width=25)
        entry.grid(row=row, column=1, sticky="w", pady=4)
        return entry

    def _add_combo(self, frame: ttk.Frame, variable: tk.StringVar, key: str, row: int):
        combo = ttk.Combobox(frame, textvariable=variable, state="readonly", width=23)
        values = self.dropdown_cache.get(key, [])
        if values:
            combo["values"] = values
            variable.set(values[0])
        else:
            combo["values"] = [""]
            variable.set("")
        combo.grid(row=row, column=1, sticky="w", pady=4)
        return combo

    @staticmethod
    def _add_result(frame: ttk.Frame, label_text: str, variable: tk.StringVar, row: int):
        ttk.Label(frame, text=label_text).grid(row=row, column=0, sticky="w", pady=6)
        ttk.Label(frame, textvariable=variable, font=("Segoe UI", 11, "bold")).grid(
            row=row, column=1, sticky="w", pady=6
        )

    def _update_pagamento_state(self, event=None):
        pagamento = self.pagamento_var.get() or ""
        if "vista" in pagamento.lower():
            self.bandeira_combo.configure(state="disabled")
            self.parcelas_combo.configure(state="disabled")
        else:
            self.bandeira_combo.configure(state="readonly")
            self.parcelas_combo.configure(state="readonly")

    def calculate(self):
        try:
            shutil.copyfile(self.excel_path, self.runtime_excel_path)
            wb = load_workbook(self.runtime_excel_path, data_only=False)
        except FileNotFoundError:
            self._set_results(None, None, None, None)
            return
        ws = wb.active
        ws["B3"] = self.contribuinte_var.get()
        ws["B4"] = parse_float(self.preco_var.get())
        ws["B5"] = parse_int(self.quantidade_var.get())
        ws["B7"] = self.pagamento_var.get()
        ws["B8"] = self.bandeira_var.get()
        ws["B9"] = self.parcelas_var.get()
        ws["B10"] = self.estado_var.get()
        wb.save(self.runtime_excel_path)
        wb.close()

        wb_result = load_workbook(self.runtime_excel_path, data_only=True)
        ws_result = wb_result.active
        frete = ws_result["B6"].value
        default_val = ws_result["B14"].value
        total_cliente = ws_result["B15"].value
        valor_unit = ws_result["I4"].value
        wb_result.close()

        self._set_results(frete, default_val, total_cliente, valor_unit)

    def _set_results(self, frete, default_val, total_cliente, valor_unit):
        self.frete_var.set(format_brl(frete))
        self.default_var.set(format_brl(default_val))
        self.total_var.set(format_brl(total_cliente))
        self.valor_unit_var.set(format_brl(valor_unit))


def main():
    root = tk.Tk()
    app = OrcamentoApp(root)
    root.mainloop()


if __name__ == "__main__":
    main()
